import os.path

import numpy as np
import pandas as pd
from openpyxl.styles import PatternFill

from utils import ROOT

DATA_PATH = ROOT / 'data/3b_all.xlsx'
OUTPUT_PATH = ROOT / f'logs/3c/3cresult.xlsx'


def normalize(arr):

    positive_arr = arr[arr > 0]
    summation = np.sum(positive_arr)
    normalized_arr = arr / summation

    return normalized_arr


table = pd.read_excel(
    DATA_PATH,
    sheet_name='Sheet1',
    header=0
)

inputs = table.iloc[:, 3:]
columns = inputs.columns
columns = columns.to_numpy()
inputs = inputs.to_numpy()

targets = table.iloc[:, 2]
targets = targets.to_numpy()

omegas, _, _, _ = np.linalg.lstsq(inputs, targets, rcond=None)

normalized_omegas = normalize(omegas)

print('Normalized alphas: ', normalized_omegas)

output_table = pd.DataFrame()
output_table['header'] = columns
output_table['normalized_omega'] = normalized_omegas

if not os.path.exists(OUTPUT_PATH.parent):
    os.mkdir(OUTPUT_PATH.parent)

xlswriter = pd.ExcelWriter(OUTPUT_PATH)
output_table.to_excel(xlswriter, sheet_name='Sheet1')
ws = xlswriter.sheets['Sheet1']

omega_column = ws['C']

for i, omega_cell in enumerate(omega_column):

    value = omega_cell.value

    if type(value) != float:
        continue

    if value > 1e-2:

        ws.cell(row=i+1, column=3).fill = PatternFill('solid', fgColor='CD2626')

    elif value > 5e-3:

        ws.cell(row=i+1, column=3).fill = PatternFill('solid', fgColor='EE7942')

    elif value < -1e-2:

        ws.cell(row=i+1, column=3).fill = PatternFill('solid', fgColor='228B22')

    elif value < -5e-3:

        ws.cell(row=i+1, column=3).fill = PatternFill('solid', fgColor='7FFF00')

xlswriter.save()
xlswriter.close()
